"""
Filtered database Excel exporter for EREBUS.

This module exports the visible EREBUS database tables to an Excel workbook.
Each exported table is written into a different worksheet and the active
execution filter is respected.

The exporter intentionally reuses DataBrowserService instead of duplicating SQL
filter logic. This keeps the exported rows consistent with the rows displayed in
the Data page.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from presentation.services.data_browser_service import DataBrowserService


class FilteredDatabaseExcelExporter:
    """
    Exports filtered EREBUS database tables to an Excel workbook.

    The exporter writes one worksheet per visible table and includes a first
    summary worksheet with export metadata.
    """

    EXPORT_PAGE_SIZE = 200
    EMPTY_VALUE = "-"

    HEADER_FILL = "1F2937"
    HEADER_FONT = "FFFFFF"
    SUMMARY_FILL = "374151"

    def __init__(self, database_service: DataBrowserService | None = None):
        """
        Initializes the exporter.

        Args:
            database_service: Optional data browser service. If omitted, the
                default EREBUS database path is used.
        """
        self.database_service = database_service or DataBrowserService()

    def export(
        self,
        output_path: str | Path,
        execution_filter: str = "",
    ) -> dict[str, Any]:
        """
        Exports visible database tables to an Excel workbook.

        Args:
            output_path: Destination .xlsx path.
            execution_filter: Optional execution identifier filter.

        Returns:
            dict: Export metadata, including exported table and row counts.

        Raises:
            FileNotFoundError: If the database does not exist.
            ValueError: If output_path is empty.
        """
        if not output_path:
            raise ValueError("Output path is required.")

        output_path = Path(output_path)

        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")

        if not self.database_service.database_exists():
            raise FileNotFoundError("EREBUS database was not found.")

        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        table_names = self.database_service.get_table_names()

        export_metadata = {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "execution_filter": execution_filter.strip(),
            "table_count": 0,
            "row_count": 0,
            "tables": [],
        }

        self._write_summary_sheet(
            workbook=workbook,
            metadata=export_metadata,
            pending=True,
        )

        used_sheet_names = {"export_summary"}

        for table_name in table_names:
            table_result = self._write_table_sheet(
                workbook=workbook,
                table_name=table_name,
                execution_filter=execution_filter,
                used_sheet_names=used_sheet_names,
            )

            export_metadata["table_count"] += 1
            export_metadata["row_count"] += table_result["row_count"]
            export_metadata["tables"].append(table_result)

        self._rewrite_summary_sheet(
            workbook=workbook,
            metadata=export_metadata,
        )

        workbook.save(output_path)

        export_metadata["output_path"] = str(output_path)
        return export_metadata

    def _write_table_sheet(
        self,
        workbook: Workbook,
        table_name: str,
        execution_filter: str,
        used_sheet_names: set[str],
    ) -> dict[str, Any]:
        """
        Writes one database table into one worksheet.

        Args:
            workbook: Target workbook.
            table_name: Database table name.
            execution_filter: Active execution filter.
            used_sheet_names: Already used Excel sheet names.

        Returns:
            dict: Table export metadata.
        """
        sheet_name = self._build_unique_sheet_name(
            table_name=table_name,
            used_sheet_names=used_sheet_names,
        )

        worksheet = workbook.create_sheet(title=sheet_name)

        columns = self.database_service.get_table_columns(table_name)
        total_rows = self.database_service.count_table_rows(
            table_name=table_name,
            execution_filter=execution_filter,
        )

        if not columns:
            worksheet.cell(row=1, column=1, value="No readable columns.")
            return {
                "table_name": table_name,
                "sheet_name": sheet_name,
                "row_count": 0,
            }

        self._write_header_row(
            worksheet=worksheet,
            columns=columns,
        )

        current_excel_row = 2
        current_page = 1

        while True:
            page_data = self.database_service.fetch_table_page(
                table_name=table_name,
                execution_filter=execution_filter,
                page=current_page,
                page_size=self.EXPORT_PAGE_SIZE,
            )

            rows = page_data.get("rows", [])

            for row in rows:
                self._write_data_row(
                    worksheet=worksheet,
                    columns=columns,
                    row=row,
                    excel_row=current_excel_row,
                )
                current_excel_row += 1

            if not page_data.get("has_next"):
                break

            current_page += 1

        if total_rows == 0:
            worksheet.cell(
                row=2,
                column=1,
                value="No rows found for the selected filter.",
            )

        self._format_table_sheet(
            worksheet=worksheet,
            columns=columns,
            total_rows=max(total_rows, 1),
        )

        return {
            "table_name": table_name,
            "sheet_name": sheet_name,
            "row_count": total_rows,
        }

    def _write_summary_sheet(
        self,
        workbook: Workbook,
        metadata: dict[str, Any],
        pending: bool = False,
    ) -> None:
        """
        Writes the export summary worksheet.

        Args:
            workbook: Target workbook.
            metadata: Export metadata.
            pending: Whether the export is still being generated.
        """
        worksheet = workbook.create_sheet(title="export_summary", index=0)

        rows = [
            ("Generated at", metadata["generated_at"]),
            (
                "Execution filter",
                metadata["execution_filter"] or "No filter applied",
            ),
            ("Status", "Generating..." if pending else "Finished"),
            ("Exported tables", metadata["table_count"]),
            ("Exported rows", metadata["row_count"]),
        ]

        for row_index, (label, value) in enumerate(rows, start=1):
            label_cell = worksheet.cell(row=row_index, column=1, value=label)
            value_cell = worksheet.cell(row=row_index, column=2, value=value)

            label_cell.font = Font(bold=True, color=self.HEADER_FONT)
            label_cell.fill = PatternFill(
                fill_type="solid",
                fgColor=self.SUMMARY_FILL,
            )
            value_cell.alignment = Alignment(wrap_text=True)

        start_row = len(rows) + 2

        worksheet.cell(row=start_row, column=1, value="Table")
        worksheet.cell(row=start_row, column=2, value="Sheet")
        worksheet.cell(row=start_row, column=3, value="Rows")

        for column in range(1, 4):
            cell = worksheet.cell(row=start_row, column=column)
            cell.font = Font(bold=True, color=self.HEADER_FONT)
            cell.fill = PatternFill(fill_type="solid", fgColor=self.HEADER_FILL)

        for index, table_info in enumerate(metadata.get("tables", []), start=start_row + 1):
            worksheet.cell(row=index, column=1, value=table_info["table_name"])
            worksheet.cell(row=index, column=2, value=table_info["sheet_name"])
            worksheet.cell(row=index, column=3, value=table_info["row_count"])

        worksheet.freeze_panes = "A7"
        self._autosize_columns(worksheet)

    def _rewrite_summary_sheet(
        self,
        workbook: Workbook,
        metadata: dict[str, Any],
    ) -> None:
        """
        Replaces the temporary summary worksheet with final metadata.

        Args:
            workbook: Target workbook.
            metadata: Final export metadata.
        """
        if "export_summary" in workbook.sheetnames:
            sheet = workbook["export_summary"]
            workbook.remove(sheet)

        self._write_summary_sheet(
            workbook=workbook,
            metadata=metadata,
            pending=False,
        )

    def _write_header_row(
        self,
        worksheet,
        columns: list[str],
    ) -> None:
        """
        Writes the header row.

        Args:
            worksheet: Target worksheet.
            columns: Column names.
        """
        for column_index, column_name in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=column_name)
            cell.font = Font(bold=True, color=self.HEADER_FONT)
            cell.fill = PatternFill(fill_type="solid", fgColor=self.HEADER_FILL)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _write_data_row(
        self,
        worksheet,
        columns: list[str],
        row: dict[str, Any],
        excel_row: int,
    ) -> None:
        """
        Writes one database row.

        Args:
            worksheet: Target worksheet.
            columns: Ordered column names.
            row: Row dictionary.
            excel_row: Excel row number.
        """
        for column_index, column_name in enumerate(columns, start=1):
            value = self._normalize_cell_value(row.get(column_name))
            cell = worksheet.cell(
                row=excel_row,
                column=column_index,
                value=value,
            )
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _format_table_sheet(
        self,
        worksheet,
        columns: list[str],
        total_rows: int,
    ) -> None:
        """
        Applies readable formatting to a table worksheet.

        Args:
            worksheet: Target worksheet.
            columns: Column names.
            total_rows: Number of exported rows.
        """
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(max(1, len(columns)))}{max(1, total_rows + 1)}"
        )
        self._autosize_columns(worksheet)

    def _autosize_columns(self, worksheet) -> None:
        """
        Autosizes worksheet columns with a reasonable maximum width.

        Args:
            worksheet: Target worksheet.
        """
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                value = cell.value

                if value is None:
                    continue

                max_length = max(max_length, len(str(value)))

            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 12),
                70,
            )

    def _normalize_cell_value(self, value: Any) -> Any:
        """
        Normalizes a database value before writing it to Excel.

        Args:
            value: Raw database value.

        Returns:
            Any: Excel-safe value.
        """
        if value is None:
            return self.EMPTY_VALUE

        if isinstance(value, bytes):
            return value.decode(errors="replace")

        if isinstance(value, bool):
            return int(value)

        return value

    def _build_unique_sheet_name(
        self,
        table_name: str,
        used_sheet_names: set[str],
    ) -> str:
        """
        Builds a valid unique Excel worksheet name.

        Excel sheet names are limited to 31 characters and cannot contain some
        special characters.

        Args:
            table_name: Desired table name.
            used_sheet_names: Already used worksheet names.

        Returns:
            str: Safe unique worksheet name.
        """
        base_name = self._sanitize_sheet_name(table_name)
        sheet_name = base_name
        suffix = 1

        while sheet_name in used_sheet_names:
            suffix_text = f"_{suffix}"
            available_length = 31 - len(suffix_text)
            sheet_name = f"{base_name[:available_length]}{suffix_text}"
            suffix += 1

        used_sheet_names.add(sheet_name)
        return sheet_name

    def _sanitize_sheet_name(self, value: str) -> str:
        """
        Sanitizes a value so it can be used as an Excel worksheet name.

        Args:
            value: Raw worksheet name.

        Returns:
            str: Safe worksheet name.
        """
        invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]

        safe_value = value.strip() or "sheet"

        for invalid_char in invalid_chars:
            safe_value = safe_value.replace(invalid_char, "_")

        return safe_value[:31]